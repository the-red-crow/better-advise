from typing import List, Dict
from academic_plan import AcademicPlan
from semester import Semester
from openpyxl import Workbook
import traceback

class ExcelExporter:
    """
    Exports academic plans to Excel format.
    """
    
    def __init__(self, output_path: str):
        """
        Initialize an ExcelExporter object.
        
        Args:
            output_path (str): Path where the Excel file will be saved
        """
        self._output_path = output_path
        self._template_path = ""
    
    def create_semester_sheet(self, semester: Semester, workbook: Workbook) -> None:
        """ Create the worksheet in the workbook for a given semester """
        new_sheet = workbook.create_sheet(str(semester.name) + " " + str(semester.year))
        new_sheet.cell(row=1,column=1,value="Course Code")
        new_sheet.cell(row=1,column=2,value="Course Name")
        new_sheet.cell(row=1,column=3,value="Credit Hours")
        new_sheet.cell(row=1,column=4,value="Taken Yet?")
        new_sheet.cell(row=1,column=5,value="Grade")
        new_sheet.cell(row=1,column=6,value="Course Description")
        iter = 2
        for course in semester.courses:
            new_sheet.cell(row=iter,column=1,value=course.code)
            new_sheet.cell(row=iter,column=2,value=course.name)
            new_sheet.cell(row=iter,column=3,value=course.getHours())
            new_sheet.cell(row=iter,column=4,value="Yes" if course.isTaken() else "No")
            new_sheet.cell(row=iter,column=5,value=course.getGrade())
            new_sheet.cell(row=iter,column=6,value=course.desc)
            iter += 1
    
    def export_academic_plan(self, plan: AcademicPlan) -> bool:
        """ Export plan to excel file """
        try:
            wb = Workbook()
            for sem in plan._semesters:
                self.create_semester_sheet(sem,wb)
            summarypage = wb.create_sheet("Summary", 0)
            for key, val in self.format_plan_summary(plan).items():
                ls = []
                ls.append(key)
                if hasattr(val, '__iter__'):
                    ls += val 
                else:
                    ls.append(val)
                summarypage.append(ls)
            wb.save(self._output_path)
            return True
        except Exception as e:
            traceback.print_exc()
            return False
        
    def format_plan_summary(self, plan: AcademicPlan) -> Dict:
        return plan.get_plan_summary()
    
    def add_prerequisite_warnings(self, issues: List[str]) -> None:
        """ Obsolete, covered by plan summary
        Add prerequisite warnings to the Excel export.
        
        Args:
            issues (List[str]): List of prerequisite issues to add
        """
        pass
    
    def save_workbook(self) -> str:
        """ Obsolete, handled by export_academic_plan, no class woorkbook variable + this function doesn't take one
        Save the workbook to file.
        
        Returns:
            str: Path to the saved file
        """
        pass