from typing import Dict, List, Tuple, Optional
from openpyxl import worksheet

from pathlib import Path
from course import Course
import openpyxl

from openpyxl.utils.exceptions import InvalidFileException
import re
class ExcelParser:
    """
    Parses Excel files to extract course schedule information.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize an ExcelParser object.
        
        Args:
            file_path (str): Path to the Excel file to parse
        """
        self._file_path = Path(file_path)

    def parse_graduate_study_plan(self, degree: str = "Software Dev") -> List[str]:
        """
        Parse a graduate study plan Excel file.
        
        Returns:
            Dict: Dictionary containing parsed graduate study plan data
        """
        if not self.validate_excel_format():
            raise InvalidFileException()

        wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        ws = wb.active
        courses = self.extract_datablock_courses(ws, degree)

        return courses

    def parse_four_year_schedule(self) -> Dict:
        """
        Parse a four-year schedule Excel file.

        Returns:
            Dict: Dictionary containing headers and row data
        """
        if not self.validate_excel_format():
            raise InvalidFileException()
        wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = None
        courses = {}

        for idx, row in enumerate(ws.iter_rows(min_row=3, max_col=27, max_row=110, values_only=True)):
            if idx == 0:
                # First row - extract headers
                headers = [cell for cell in row if cell is not None]
            else:
                # Row not empty
                if row[0]:
                    available_semster = []
                    for i, semester in enumerate(row[2:], start=2):
                        if semester:
                            match = re.search(r'D|N|O', semester)
                            if match:
                                available_semster.append(headers[i])
                    courses[row[0]] = available_semster

        wb.close()
        return courses


    def validate_excel_format(self) -> bool:
        """
        Validate the Excel file format.

        Returns:
            bool: True if valid format, False otherwise
        """
        supported_ext = {'.xlsx', '.xlsm', '.xltx', '.xltm'}
        if self._file_path.suffix.lower() not in supported_ext:
            return False

        try:
            openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        except (InvalidFileException, ValueError, OSError):
            return False

        return True

    def find_datablock(self, ws: worksheet, partial_match: str) -> Optional[Tuple[int, str]]:
        """
        Find a datablock (degree program) based on partial text match.

        Args:
            ws worksheet: The worksheet to search
            partial_match (str): Partial text to search for in datablock headers

        Returns:
            Optional[Tuple[int, str]]: Tuple of (row_index, full_datablock_name) if found,
                                       None otherwise
        """
        search_term = partial_match.lower()

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for cell_value in row:
                if cell_value is None:
                    continue

                cell_str = str(cell_value).strip()
                compare_str = cell_str.lower()

                # Check if this looks like a datablock header (contains "-")
                if " - " in cell_str and search_term in compare_str:
                    return (row_idx, cell_str)

        return None

    def extract_datablock_courses(self, ws: worksheet, partial_match: str) -> List[str]:
        """
        Extract courses from a datablock's Fall section (3 rows x 5 columns starting
        from the cell in the bottom right direction of the datablock header).

        Args:
            ws worksheet: The worksheet to search
            partial_match (str): Partial text to search for in datablock headers

        Returns:
            List[str]: List of courses found in the Fall section
        """
        datablock_result = self.find_datablock(ws, partial_match)
        if not datablock_result:
            return []

        header_row, _ = datablock_result
        courses = []

        header_col = 3  # Default column C (index 3)
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value and " - " in str(cell_value):
                header_col = col
                break

        # Start collecting courses 1 row below header, from column right of header
        start_row = header_row + 1
        start_col = header_col + 1

        # Extract 3 rows x 5 columns
        for row_offset in range(3):
            current_row = start_row + row_offset
            for col_offset in range(5):
                cell_value = ws.cell(row=current_row, column=start_col + col_offset).value
                if cell_value:
                    course_str = str(cell_value).strip()
                    # Skip empty cells and section headers
                    if course_str and course_str.lower() not in ["elective"]:
                        courses.append(course_str)

        return courses

if __name__ == "__main__":
    test = ExcelParser("input/Graduate Study Plans -revised.xlsx")
    test.parse_graduate_study_plan()
    test4 = ExcelParser("input/4-year schedule.xlsx")
    test4.parse_four_year_schedule()
    pass